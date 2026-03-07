"""
Build the Task 1 Excel workbook: Global Market Value of Coal vs Oil vs Gas
Data year: 2024 (primary), 2023 (where 2024 unavailable)
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
import csv
import os

wb = openpyxl.Workbook()

# Style definitions
header_font = Font(bold=True, size=11)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font_white = Font(bold=True, size=11, color="FFFFFF")
section_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
number_fmt = '#,##0'
number_fmt_1d = '#,##0.0'
money_fmt = '#,##0.00'
pct_fmt = '0.0%'
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def style_header(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border

def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 40)

# ============================================================
# SHEET 1: COAL
# ============================================================
ws_coal = wb.active
ws_coal.title = "Coal"

# Header row
coal_headers = [
    "Rank", "Country", "Coal Type", "Production 2024 (Mt)",
    "Exports (Mt)", "Imports (Mt)", "Domestic Consumption (Mt)",
    "Price ($/t)", "Price Benchmark", "Price Type",
    "Estimated Market Value ($M)", "Data Year", "Source ID", "Notes"
]
for col, h in enumerate(coal_headers, 1):
    ws_coal.cell(row=1, column=col, value=h)
style_header(ws_coal, 1, len(coal_headers))

# Coal data - 2024 figures, split by type where available
coal_data = [
    # Rank, Country, Type, Prod Mt, Exports, Imports, DomCons, Price, Benchmark, PriceType, MktVal, Year, SrcID, Notes
    [1, "China", "Thermal", 3710, None, 450, 3260, 107, "Qinhuangdao 5500 kcal", "Delivered (port)", 396970, 2024, "C1,C2", "~80% of production; QHD avg ~RMB 750-863/t; IEA 4,666 Mt total"],
    [1, "China", "Metallurgical", 465, None, 93, 372, 242, "PLV HCC FOB Aus proxy", "Import parity", 112530, 2024, "C1,C3", "Coking coal; import price used as proxy"],
    [1, "China", "Other/Lignite", 491, None, None, 491, 50, "Estimated", "Mine-gate", 24550, 2024, "C1", "Anthracite, PCI, other grades"],
    [2, "India", "Thermal", 1020, None, 180, 1020, 20, "Coal India notified avg", "Mine-gate", 20400, 2024, "C4,C5", "CIL FSA ~Rs 1450/t; FY24-25; IEA 1,089 Mt (incl. met)"],
    [2, "India", "Metallurgical", 28, None, 65, 28, 242, "PLV HCC import parity", "CIF", 6776, 2024, "C4", "Small domestic met coal; mostly imported"],
    [3, "Indonesia", "Thermal", 833, 556, None, 277, 122, "HBA 6322 kcal GAR", "FOB", 101626, 2024, "C6,C7", "Record 831-836 Mt; HBA avg ~$122/t; DMO domestic price capped"],
    [4, "USA", "Thermal", 400, 45, None, 355, 42, "EIA avg thermal mine price", "Mine-gate", 16800, 2024, "C8", "512 MMst total = ~465 Mt; $37.85/short ton = ~$42/t metric"],
    [4, "USA", "Metallurgical", 65, 52, None, 13, 198, "EIA avg met mine price", "Mine-gate", 12870, 2024, "C8", "$180/short ton = ~$198/t metric"],
    [5, "Australia", "Thermal", 250, 210, None, 40, 134, "Newcastle FOB 6000 kcal", "FOB", 33500, 2024, "C9,C10", "~463 Mt saleable total; ~55% thermal"],
    [5, "Australia", "Metallurgical", 213, 153, None, 60, 242, "PLV HCC FOB", "FOB", 51546, 2024, "C9,C10", "World's largest met coal exporter"],
    [6, "Russia", "Thermal", 360, 170, None, 190, 80, "Estimated blended", "Mixed FOB/domestic", 28800, 2024, "C11", "~427 Mt total; sanctions impact; exports -6-8%; domestic ~$30-50/t"],
    [6, "Russia", "Metallurgical", 67, 28, None, 39, 200, "Estimated", "FOB", 13400, 2024, "C11", "Some met coal exports via Pacific ports"],
    [7, "South Africa", "Thermal", 237, 52, None, 185, 88, "Richards Bay 5500 kcal", "FOB", 20856, 2024, "C12,C13", "~235-240 Mt; RBCT exports 52.1 Mt"],
    [8, "Kazakhstan", "Mixed", 109, 32, None, 77, 60, "Estimated", "Mine-gate", 6540, 2024, "C14", "108.5 Mt; mix thermal + met"],
    [9, "Mongolia", "Metallurgical", 75, 56, None, 19, 200, "China import price proxy", "FOB border", 15000, 2024, "C15", "~98 Mt total; mostly met for China; +20%"],
    [9, "Mongolia", "Thermal", 23, 28, None, None, 80, "Estimated", "FOB border", 1840, 2024, "C15", "Some thermal exports"],
    [10, "Germany", "Lignite", 86, None, None, 86, 25, "Estimated pit-mouth", "Mine-gate", 2150, 2024, "C16", "~80-92 Mt; 100% lignite; hard coal mining ceased 2018"],
    [11, "Turkey", "Lignite", 87, None, 25, 87, 30, "Estimated", "Mine-gate", 2610, 2024, "C17", "+17% growth; >90% lignite"],
    [12, "Poland", "Hard coal", 44, None, 15, 44, 85, "Estimated", "Mine-gate", 3740, 2024, "C18", "Declining; committed close by 2049"],
    [12, "Poland", "Lignite", 40, None, None, 40, 25, "Estimated pit-mouth", "Mine-gate", 1000, 2024, "C18", "Belchatow + Turow mines"],
    [13, "Colombia", "Thermal", 63, 58, None, 5, 88, "Richards Bay proxy", "FOB", 5544, 2024, "C19", "Almost entirely thermal for export"],
    [14, "Vietnam", "Thermal", 44, None, 40, 44, 80, "Estimated", "Mixed", 3520, 2024, "C20", "43.8 Mt; -2.7%; anthracite from Quang Ninh"],
    [15, "Canada", "Metallurgical", 29, 25, None, 4, 242, "PLV HCC FOB", "FOB", 7018, 2024, "C21", "42.6 Mt total; 67% met"],
    [15, "Canada", "Thermal", 14, None, None, 14, 60, "Estimated", "Mine-gate", 840, 2024, "C21", "Thermal declining"],
    [16, "Czech Republic", "Lignite", 24, None, None, 24, 25, "Estimated", "Mine-gate", 600, 2024, "C22", "~25 Mt total; -17%"],
    [16, "Czech Republic", "Hard coal", 1, None, None, 1, 80, "Estimated", "Mine-gate", 80, 2024, "C22", "Minimal hard coal"],
    [17, "Ukraine", "Mixed", 17, None, None, 17, 70, "Estimated", "Mine-gate", 1190, 2024, "C23", "17.4 Mt; -12%; war-impacted"],
    [18, "Pakistan", "Thermal/Lignite", 17, None, 20, 17, 30, "Estimated", "Mine-gate", 510, 2024, "C24", "~17 Mt; growing from Thar lignite"],
    [19, "Mozambique", "Mixed", 16, 14, None, 2, 130, "FOB estimate", "FOB", 2080, 2024, "C25", "16.3 Mt; +9.4%; Moatize mine thermal + met"],
    [20, "Philippines", "Thermal", 17, None, 25, 17, 60, "Estimated", "Mine-gate", 1020, 2024, "C26", "16.5 Mt; Semirara Island"],
    [21, "Botswana", "Thermal", 3, 1, None, 2, 50, "Estimated", "Mine-gate", 150, 2024, "C27", "~2-3 Mt; Morupule Coal Mine; sub-bituminous"],
    [None, "Rest of World", "Mixed", 252, None, None, 252, 60, "Estimated (blended)", "Various", 15120, 2024, "C13", "~60+ smaller producing countries; IEA global total 9,100 Mt"],
]

for r, row in enumerate(coal_data, 2):
    for c, val in enumerate(row, 1):
        cell = ws_coal.cell(row=r, column=c, value=val)
        cell.border = thin_border
        if c in [4, 5, 6, 7, 11]:  # numeric columns
            cell.number_format = number_fmt
        if c == 8:  # price
            cell.number_format = money_fmt

# Summary row
summary_row = len(coal_data) + 3
ws_coal.cell(row=summary_row, column=1, value="TOTAL").font = header_font
ws_coal.cell(row=summary_row, column=4, value="~9,100 Mt (IEA Coal 2025)").font = header_font
ws_coal.cell(row=summary_row, column=11, value="~$900B+ (est.)").font = header_font
ws_coal.cell(row=summary_row, column=14, value="Sum of country estimates; actual total depends on domestic price assumptions for non-reported markets")

auto_width(ws_coal)

# ============================================================
# SHEET 2: OIL
# ============================================================
ws_oil = wb.create_sheet("Oil")

oil_headers = [
    "Rank", "Country", "Production 2023 (kb/d)", "Production (M bbl/yr)",
    "Primary Crude Grade", "API Gravity", "Sulfur %", "Quality Class",
    "Country Benchmark Price ($/bbl)", "Differential to Brent ($/bbl)",
    "Pricing Basis", "Estimated Market Value ($B)",
    "Confidence", "Data Year", "Source ID", "Notes"
]
for col, h in enumerate(oil_headers, 1):
    ws_oil.cell(row=1, column=col, value=h)
style_header(ws_oil, 1, len(oil_headers))

# Read oil CSV if available, otherwise use research data
oil_csv_path = os.path.join(os.path.dirname(__file__), "oil_country_benchmark_mapping.csv")
oil_row = 2

# Top 20 + key additional countries from research
oil_data = [
    [1, "United States", 12927, 4718, "WTI (Cushing)", 39.6, 0.24, "Light Sweet", 77.58, -4.91, "WTI", 366.1, "High", 2023, "O1,O2", "Largest producer; 20+ mb/d in 2024"],
    [2, "Russia", 10300, 3760, "Urals", 31.7, 1.30, "Medium Sour", 63.00, -20.70, "Brent (discounted)", 236.8, "Medium", 2023, "O3,O4", "Sanctions discount; ESPO ~$73 to China"],
    [3, "Saudi Arabia", 9612, 3509, "Arab Light", 32.8, 1.77, "Medium Sour", 80.50, -2.00, "Dubai/Oman", 282.4, "Medium", 2023, "O5", "OPEC+ voluntary cuts; fell to 9.2 mb/d in 2024"],
    [4, "Canada", 4832, 1764, "WCS / Syncrude / Conv Light", 20.5, 3.50, "Heavy Sour (WCS)", 65.00, -17.49, "WTI", 114.6, "High", 2023, "O6", "Weighted avg: WCS $59, Syncrude $77, conv $77"],
    [5, "Iraq", 4420, 1613, "Basra Medium", 27.0, 2.80, "Medium Sour", 78.00, -4.49, "Dubai/Oman", 125.8, "Medium", 2023, "O5", ""],
    [6, "China", 4111, 1501, "Daqing / Shengli", 33.0, 0.10, "Medium Sweet", 82.00, -0.49, "Dubai (import)", 123.0, "Low", 2023, "O7", "Domestic production priced at import parity"],
    [7, "UAE", 3530, 1289, "Murban", 40.0, 0.78, "Light Sour", 80.00, -2.49, "Murban (IFAD)", 103.1, "Medium", 2023, "O5,O8", "ICE Futures Abu Dhabi since 2021"],
    [8, "Iran", 3228, 1178, "Iran Heavy", 31.0, 1.70, "Medium Sour", 68.00, -14.49, "Brent (discounted)", 80.1, "Low", 2023, "O9", "Sanctions; realized ~$65-70; official ~$81"],
    [9, "Brazil", 3410, 1245, "Tupi/Lula", 28.4, 0.32, "Medium Sweet", 80.50, -1.99, "Brent", 100.2, "Medium", 2023, "O10", "Pre-salt deepwater"],
    [10, "Kuwait", 2676, 977, "Kuwait Export (KEC)", 30.2, 2.72, "Medium Sour", 79.00, -3.49, "Dubai/Oman", 77.2, "Medium", 2023, "O5", ""],
    [11, "Norway", 1901, 694, "Johan Sverdrup / Ekofisk", 28.0, 0.20, "Medium Sweet", 82.49, 0.00, "Brent", 57.2, "High", 2023, "O1", "Part of Brent basket"],
    [12, "Mexico", 1863, 680, "Maya", 22.0, 3.50, "Heavy Sour", 72.00, -10.49, "WTI (Gulf)", 49.0, "Medium", 2023, "O11", "Pemex data; declining production"],
    [13, "Kazakhstan", 1801, 657, "CPC Blend / Tengiz", 45.3, 0.56, "Light Sweet", 79.00, -3.49, "Brent (Black Sea)", 51.9, "Medium", 2023, "O12", ""],
    [14, "Qatar", 1748, 638, "Qatar Marine", 35.8, 1.47, "Medium Sour", 80.00, -2.49, "Dubai/Oman", 51.0, "Medium", 2023, "O5", ""],
    [15, "Algeria", 1474, 538, "Saharan Blend", 45.5, 0.05, "Light Sweet", 83.00, 0.51, "Brent", 44.7, "Medium", 2023, "O5", "Premium to Brent for quality"],
    [16, "Nigeria", 1250, 456, "Bonny Light", 35.4, 0.14, "Light Sweet", 83.50, 1.01, "Brent", 38.1, "Medium", 2023, "O5", "Range $74-$98 in 2023"],
    [17, "Angola", 1143, 417, "Girassol", 30.5, 0.34, "Medium Sweet", 81.00, -1.49, "Brent", 33.8, "Medium", 2023, "O5", ""],
    [18, "Libya", 1180, 431, "Es Sider", 37.0, 0.44, "Light Sweet", 82.50, 0.01, "Brent", 35.5, "Medium", 2023, "O5", "Tracks Brent closely"],
    [19, "Oman", 1064, 388, "Oman Crude", 31.4, 2.00, "Medium Sour", 82.00, -0.49, "Dubai/Oman", 31.8, "Medium", 2023, "O5", "Dubai/Oman benchmark"],
    [20, "UK", 670, 245, "Brent Blend", 38.3, 0.37, "Light Sweet", 82.49, 0.00, "Brent", 20.2, "High", 2023, "O1", "Declining North Sea"],
    [21, "Guyana", 400, 146, "Liza", 31.9, 0.59, "Medium Sweet", 81.00, -1.49, "Brent", 11.8, "Low", 2023, "O13", "Fastest growing producer"],
    [22, "Colombia", 750, 274, "Vasconia / Castilla", 24.0, 1.00, "Medium Sour", 78.00, -4.49, "Brent", 21.4, "Low", 2023, "O14", ""],
    [23, "Azerbaijan", 620, 226, "Azeri Light (BTC)", 35.3, 0.14, "Light Sweet", 82.00, -0.49, "Brent", 18.5, "Medium", 2023, "O15", ""],
    [24, "Ecuador", 475, 173, "Oriente", 24.0, 1.50, "Heavy Sour", 74.00, -8.49, "WTI", 12.8, "Low", 2023, "O16", ""],
    [25, "Venezuela", 750, 274, "Merey 16", 16.0, 2.60, "Heavy Sour", 64.37, -18.12, "Brent (discounted)", 17.6, "Low", 2023, "O5", "Sanctions; opaque pricing"],
]

for r, row in enumerate(oil_data, 2):
    for c, val in enumerate(row, 1):
        cell = ws_oil.cell(row=r, column=c, value=val)
        cell.border = thin_border
        if c in [3, 4]:
            cell.number_format = number_fmt
        if c in [9, 10, 12]:
            cell.number_format = money_fmt
        if c in [6, 7]:
            cell.number_format = number_fmt_1d

# Oil summary
oil_summary_row = len(oil_data) + 3
ws_oil.cell(row=oil_summary_row, column=1, value="SUMMARY").font = header_font
ws_oil.cell(row=oil_summary_row, column=3, value="~81,800 kb/d").font = header_font
ws_oil.cell(row=oil_summary_row, column=12, value="$2,270B (country-specific) / $2,460B (flat Brent)").font = header_font
ws_oil.cell(row=oil_summary_row + 1, column=1, value="Notes:")
ws_oil.cell(row=oil_summary_row + 1, column=2, value="Country-specific pricing reduces total by ~$190B (7.7%) vs flat Brent. Driven by Russia (-$160B), Canada (-$31B), Iran (-$17B), Mexico (-$7B)")
ws_oil.cell(row=oil_summary_row + 2, column=2, value="Full 80-country mapping in oil_country_benchmark_mapping.csv. Remaining ~60 countries (~19% of production) use Brent/Dubai proxy with quality adjustments.")

auto_width(ws_oil)

# ============================================================
# SHEET 3: GAS
# ============================================================
ws_gas = wb.create_sheet("Gas")

gas_headers = [
    "Rank", "Country", "Production (bcm)", "Domestic Consumption (bcm)",
    "LNG Exports (Mt)", "Pipeline Exports (bcm)",
    "Pricing Region", "Benchmark Price ($/MMBtu)", "Est. Domestic Price ($/MMBtu)",
    "Price Regime", "Estimated Market Value ($M)",
    "Data Year", "Source ID", "Notes"
]
for col, h in enumerate(gas_headers, 1):
    ws_gas.cell(row=1, column=col, value=h)
style_header(ws_gas, 1, len(gas_headers))

gas_data = [
    [1, "United States", 1035, 902, 84.5, None, "North America", 2.57, 2.57, "Henry Hub", 90790, 2023, "G1,G2", "25% of global; largest LNG exporter"],
    [2, "Russia", 586, 477, 31.4, 100, "Regulated", 1.80, 1.80, "Regulated (Gazprom)", 35997, 2023, "G1,G3", "Domestic regulated ~5500 RUB/tcm"],
    [3, "Iran", 259, 245, None, 15, "Subsidized", 0.50, 0.50, "Multi-tier tariff", 4419, 2023, "G1,G4", "World's cheapest gas; $42B+ implicit subsidy"],
    [4, "China", 234, 434, None, None, "Mixed", 8.50, 8.50, "Mixed import+domestic", 67881, 2023, "G1,G5", "Net importer; consumption >> production"],
    [5, "Canada", 190, 129, None, None, "North America", 2.14, 2.14, "AECO-C", 13873, 2023, "G1,G6", "AECO-C extremely depressed"],
    [6, "Qatar", 178, 40, 78.2, None, "Subsidized/LNG", 1.25, 1.25, "QP-administered", 7589, 2023, "G1,G7", "#3 LNG exporter; domestic very cheap"],
    [7, "Australia", 152, 45, 79.6, None, "Export netback", 9.00, 9.00, "LNG netback", 46694, 2023, "G1,G8", "#2 LNG exporter"],
    [8, "Saudi Arabia", 121, 122, None, None, "Administered", 1.25, 1.25, "Government-set", 5160, 2023, "G1,G4", "Set at $1.25/MMBtu since 2016"],
    [9, "Norway", 116, 5, 2.5, 108, "Europe", 13.00, 13.00, "TTF", 51464, 2023, "G1,G9", "Almost all exported via pipeline to Europe"],
    [10, "Algeria", 102, 51, 13.0, 38, "Subsidized/export", 1.00, 1.00, "Subsidized domestic", 3479, 2023, "G1,G10", "Dual pricing; exports at int'l, domestic <$1"],
    [11, "Turkmenistan", 87, 35, None, 52, "Regulated", 1.00, 1.00, "Regulated", 2968, 2023, "G1,G11", "Pipeline exports to China"],
    [12, "Malaysia", 76, 40, 26.8, None, "Mixed/LNG", 6.00, 6.00, "Mixed", 15556, 2023, "G1,G12", "#5 LNG exporter"],
    [13, "Egypt", 65, 60, 7.0, None, "Subsidized", 2.50, 2.50, "Gradually deregulating", 5544, 2023, "G1,G13", "Reforms raising prices"],
    [14, "Indonesia", 57, 47, 15.6, None, "Regulated/LNG", 5.00, 5.00, "Regulated", 9724, 2023, "G1,G14", "#6 LNG exporter"],
    [15, "UAE", 57, 71, 5.0, None, "Administered", 1.50, 1.50, "Old contracts", 2916, 2023, "G1,G4", "Net gas importer; Dolphin pipeline"],
    [16, "Uzbekistan", 48, 55, None, None, "Regulated", 1.00, 1.00, "Regulated", 1638, 2023, "G1,G11", "Net importer; gradually reforming"],
    [17, "Argentina", 43, 43, None, None, "Regulated", 4.50, 4.50, "Regulated", 6604, 2023, "G1,G15", "Vaca Muerta shale growth"],
    [18, "Oman", 41, 41, 11.4, None, "Administered", 3.39, 3.39, "Reformed administered", 4744, 2023, "G1,G4", "Raised from $1.50 to $3.39"],
    [19, "Nigeria", 40, 40, 13.0, None, "Subsidized/LNG", 2.00, 2.00, "Dual pricing", 2730, 2023, "G1,G16", "Low domestic; LNG at int'l"],
    [20, "United Kingdom", 38, 62, None, None, "Europe", 12.50, 12.50, "NBP", 16210, 2023, "G1,G17", "Net importer; tracks TTF"],
    [21, "Pakistan", 37, 37, None, None, "Regulated", 3.50, 3.50, "Regulated", 4418, 2023, "G1,G18", "Declining production"],
    [22, "Azerbaijan", 34, 12, None, 22, "Mixed", 3.00, 3.00, "Mixed", 3480, 2023, "G1,G19", "SGC pipeline to Europe"],
    [23, "India", 33, 70, None, None, "Regulated", 6.50, 6.50, "APM + import", 7321, 2023, "G1,G20", "Large net importer"],
    [24, "Mexico", 31, 100, None, None, "HH-linked", 3.50, 3.50, "HH-linked", 3702, 2023, "G1,G21", "Large net importer from US"],
    [25, "Thailand", 31, 48, None, None, "Regulated", 7.00, 7.00, "Regulated", 7405, 2023, "G1,G22", "Net importer (Myanmar pipeline + LNG)"],
    [26, "Trinidad & Tobago", 28, 18, 10.0, None, "LNG export", 5.00, 5.00, "LNG export", 4775, 2023, "G1,G23", "#10 LNG exporter"],
    [27, "Kazakhstan", 27, 17, None, None, "Regulated", 1.50, 1.50, "Regulated", 1382, 2023, "G1,G11", "Low domestic"],
    [28, "Bangladesh", 25, 25, None, None, "Regulated", 2.00, 2.00, "Regulated", 1706, 2023, "G1,G24", "Declining; no exports"],
    [29, "Israel", 23, 15, None, 8, "Mixed", 5.50, 5.50, "Leviathan/Tamar contracts", 4317, 2023, "G1,G25", "Pipeline exports to Egypt, Jordan"],
    [30, "Brazil", 23, 35, None, None, "Regulated", 6.00, 6.00, "Regulated", 4710, 2023, "G1,G26", "Pre-salt associated gas"],
]

for r, row in enumerate(gas_data, 2):
    for c, val in enumerate(row, 1):
        cell = ws_gas.cell(row=r, column=c, value=val)
        cell.border = thin_border
        if c in [3, 4, 5, 6, 11]:
            cell.number_format = number_fmt
        if c in [8, 9]:
            cell.number_format = money_fmt

gas_summary_row = len(gas_data) + 3
ws_gas.cell(row=gas_summary_row, column=1, value="GLOBAL TOTAL").font = header_font
ws_gas.cell(row=gas_summary_row, column=3, value="~4,075 bcm (2023) / 4,124 bcm (2024)").font = header_font
ws_gas.cell(row=gas_summary_row, column=11, value="~$670-680B").font = header_font
ws_gas.cell(row=gas_summary_row + 1, column=2, value="94 producing countries total. Top 30 shown above = ~90% of global. Full list in gas_production_by_country_2023.csv")
ws_gas.cell(row=gas_summary_row + 2, column=2, value="~31% of gas consumed at <$2/MMBtu (subsidized); ~60% at market prices (>$5/MMBtu). IGU global weighted avg: $4.88/MMBtu")

auto_width(ws_gas)

# ============================================================
# SHEET 4: TRANSPORT COSTS
# ============================================================
ws_transport = wb.create_sheet("Transport Costs")

# Coal seaborne
ws_transport.cell(row=1, column=1, value="COAL SEABORNE FREIGHT").font = Font(bold=True, size=13)
coal_transport_headers = ["Route", "Origin", "Destination", "Vessel Type", "Cost ($/t)", "Cost Range", "Data Year", "Source ID"]
for col, h in enumerate(coal_transport_headers, 1):
    ws_transport.cell(row=2, column=col, value=h)
style_header(ws_transport, 2, len(coal_transport_headers))

coal_freight = [
    ["Australia-Japan", "Newcastle", "Japan", "Panamax", 14.50, "$10-20", 2024, "T1"],
    ["South Africa-Rotterdam", "Richards Bay", "Rotterdam", "Capesize", 9.80, "$7-14", 2024, "T1"],
    ["Indonesia-South China", "Kalimantan", "South China", "Panamax", 8.10, "$6-12", 2024, "T1"],
    ["Indonesia-India West", "Kalimantan", "West India", "Supramax", 14.00, "$10-18", 2024, "T1"],
    ["Indonesia-India East", "Kalimantan", "East India", "Panamax", 7.00, "$5-10", 2024, "T1"],
    ["Russia-North China", "Vostochny", "North China", "Panamax", 7.00, "$5-10", 2023, "T2"],
    ["Colombia-Europe", "Puerto Bolivar", "ARA", "Capesize", 12.00, "$8-16", 2024, "T1"],
    ["USA-Europe", "Hampton Roads", "ARA", "Capesize", 15.00, "$10-20", 2024, "T1"],
    ["Mozambique-India", "Nacala", "India", "Panamax", 18.00, "$14-22", 2024, "T1"],
]
for r, row in enumerate(coal_freight, 3):
    for c, val in enumerate(row, 1):
        ws_transport.cell(row=r, column=c, value=val).border = thin_border

# Coal inland
r_start = len(coal_freight) + 5
ws_transport.cell(row=r_start, column=1, value="COAL INLAND TRANSPORT").font = Font(bold=True, size=13)
inland_headers = ["Country", "Route", "Mode", "Distance (km)", "Cost ($/t)", "Cost Range", "Source ID"]
for col, h in enumerate(inland_headers, 1):
    ws_transport.cell(row=r_start + 1, column=col, value=h)
style_header(ws_transport, r_start + 1, len(inland_headers))

coal_inland = [
    ["China", "Shanxi to Qinhuangdao", "Rail", 653, 12, "$11-35 total pit-to-port", "T3"],
    ["India", "Average mine to plant", "Rail", 586, 9, "$7-11", "T4"],
    ["USA", "CAPP to export terminal", "Rail", 500, 38, "$35-40/short ton", "T5"],
    ["USA", "PRB to Gulf ports", "Rail", 2400, 35, "$25-44/short ton", "T5"],
    ["Australia", "Hunter Valley to Newcastle", "Rail", 200, 7, "A$5-15/t", "T6"],
    ["Indonesia", "Kalimantan barge", "Barge", 100, 4, "$3-5 short haul", "T7"],
    ["South Africa", "Mpumalanga to Richards Bay", "Rail", 588, 6, "$5-7", "T8"],
    ["Russia", "Kuzbass to Vostochny", "Rail", 5500, 50, "$45-55+", "T9"],
]
for r_off, row in enumerate(coal_inland, 2):
    for c, val in enumerate(row, 1):
        ws_transport.cell(row=r_start + r_off, column=c, value=val).border = thin_border

# Oil tanker rates
r_oil = r_start + len(coal_inland) + 4
ws_transport.cell(row=r_oil, column=1, value="OIL TANKER RATES").font = Font(bold=True, size=13)
oil_t_headers = ["Route", "Origin", "Destination", "Vessel Type", "Cost ($/bbl)", "Cost Range", "Source ID"]
for col, h in enumerate(oil_t_headers, 1):
    ws_transport.cell(row=r_oil + 1, column=col, value=h)
style_header(ws_transport, r_oil + 1, len(oil_t_headers))

oil_tanker = [
    ["ME-Asia", "AG", "China/Japan", "VLCC", 2.25, "$1.50-3.00", "T10"],
    ["USG-China", "US Gulf", "China", "VLCC", 4.50, "$4.00-5.00", "T10"],
    ["USG-Europe", "US Gulf", "NW Europe", "VLCC/Suezmax", 2.50, "$2.00-3.00", "T10"],
    ["WAF-Europe", "W Africa", "NW Europe", "Suezmax", 3.50, "$2.50-4.50", "T10"],
    ["WAF-Asia", "W Africa", "China", "VLCC", 3.00, "$2.50-4.00", "T10"],
    ["Russia-India", "Baltic/Black Sea", "India", "Aframax", 5.00, "$4.00-7.00", "T11"],
]
for r_off, row in enumerate(oil_tanker, 2):
    for c, val in enumerate(row, 1):
        ws_transport.cell(row=r_oil + r_off, column=c, value=val).border = thin_border

# LNG chain
r_lng = r_oil + len(oil_tanker) + 4
ws_transport.cell(row=r_lng, column=1, value="LNG VALUE CHAIN COSTS ($/MMBtu)").font = Font(bold=True, size=13)
lng_headers = ["Corridor", "Wellhead", "Liquefaction", "Shipping", "Regasification", "Total Landed", "Source ID"]
for col, h in enumerate(lng_headers, 1):
    ws_transport.cell(row=r_lng + 1, column=col, value=h)
style_header(ws_transport, r_lng + 1, len(lng_headers))

lng_chain = [
    ["US -> Europe", 2.57, 3.00, 0.51, 0.40, 6.50, "T12,T13"],
    ["US -> Asia", 2.57, 3.00, 1.08, 0.40, 7.05, "T12,T13"],
    ["Qatar -> Asia", 1.25, 1.75, 0.70, 0.40, 4.10, "T12,T14"],
    ["Australia -> Asia", 3.50, 2.50, 0.45, 0.40, 6.85, "T12,T15"],
]
for r_off, row in enumerate(lng_chain, 2):
    for c, val in enumerate(row, 1):
        ws_transport.cell(row=r_lng + r_off, column=c, value=val).border = thin_border

auto_width(ws_transport)

# ============================================================
# SHEET 5: SUMMARY
# ============================================================
ws_summary = wb.create_sheet("Summary")

ws_summary.cell(row=1, column=1, value="GLOBAL FOSSIL FUEL MARKET VALUE SUMMARY").font = Font(bold=True, size=14)
ws_summary.cell(row=2, column=1, value="Data Year: 2024 (coal, gas) / 2023 (oil)").font = Font(italic=True)

summary_headers = ["Commodity", "Global Production", "Unit", "Reference Price", "Price Unit",
                    "Market Value (Low)", "Market Value (High)", "Recommended Estimate", "Key Methodology Note"]
for col, h in enumerate(summary_headers, 1):
    ws_summary.cell(row=4, column=col, value=h)
style_header(ws_summary, 4, len(summary_headers))

summary_data = [
    ["Coal (total)", "9,100 Mt", "Mt", "Various", "$/t", "$500B", "$900B", "$700-800B",
     "IEA Coal 2025: 9.1 Bt production (2024 record). Demand: 8.77 Bt. Wide range due to domestic pricing uncertainty; ~82% consumed domestically"],
    ["  Thermal coal", "~7,650 Mt", "Mt", "$88-134", "$/t FOB", None, None, None,
     "Newcastle $134/t, RB $88/t, HBA $122/t, China domestic ~$107/t, India ~$20/t"],
    ["  Met coal", "~1,050 Mt", "Mt", "$242", "$/t FOB", None, None, None,
     "PLV HCC FOB Australia avg $242/t in 2024"],
    ["  Lignite", "~450 Mt", "Mt", "$20-30", "$/t", None, None, None,
     "Low value; Germany, Turkey, Poland, Czech Rep; mined and burned on-site"],
    ["Oil (crude+condensate)", "81,800 kb/d", "kb/d", "$80.56 (Brent)", "$/bbl", "$2,270B", "$2,460B", "$2,270B",
     "Low = country-specific pricing; High = flat Brent; Russia sanctions -$160B, Canada WCS -$31B"],
    ["Oil (incl. NGLs)", "97,000 kb/d", "kb/d", "$80.56 (Brent)", "$/bbl", "$2,440B", "$2,850B", "$2,500-2,700B",
     "NGLs trade at 30-60% of crude; using Brent overstates NGL value"],
    ["Natural Gas", "4,124 bcm", "bcm", "$4.88 (global avg)", "$/MMBtu", "$670B", "$680B", "$670-680B",
     "IGU weighted avg; 31% consumed at <$2 (subsidized); if all at market prices would be ~$1.2T"],
    ["TOTAL FOSSIL FUELS", None, None, None, None, "$3,440B", "$3,940B", "$3,650-4,000B",
     "Sum of recommended estimates; ~$3.7T central estimate"],
]

for r, row in enumerate(summary_data, 5):
    for c, val in enumerate(row, 1):
        cell = ws_summary.cell(row=r, column=c, value=val)
        cell.border = thin_border
        if "TOTAL" in str(row[0]):
            cell.font = header_font

# Transport cost as % of value
ws_summary.cell(row=14, column=1, value="").font = header_font
ws_summary.cell(row=15, column=1, value="TRANSPORT COST AS % OF COMMODITY VALUE").font = Font(bold=True, size=13)
tc_headers = ["Commodity", "Transport Cost Range", "As % of Commodity Value", "Notes"]
for col, h in enumerate(tc_headers, 1):
    ws_summary.cell(row=16, column=col, value=h)
style_header(ws_summary, 16, len(tc_headers))

tc_data = [
    ["Crude Oil", "$1.50-5.00/bbl", "2-7%", "Highly efficient VLCCs; oil is high value per unit weight"],
    ["Thermal Coal (seaborne)", "$7-20/t", "5-25%", "Low value commodity; distance matters enormously"],
    ["Thermal Coal (Russia inland)", "$45-55/t", "50-90%", "Extreme case: Kuzbass to Pacific 5,500 km; logistics ~90% of price"],
    ["LNG (total chain)", "$3-7/MMBtu", "10-30%", "Includes liquefaction + shipping + regas; Qatar cheapest ($4.10 landed)"],
    ["Pipeline Gas", "$0.50-2.50/1000km", "10-40%", "Distance-dependent; LNG competitive beyond 3,000-5,000 km"],
]
for r, row in enumerate(tc_data, 17):
    for c, val in enumerate(row, 1):
        ws_summary.cell(row=r, column=c, value=val).border = thin_border

auto_width(ws_summary)

# ============================================================
# SHEET 6: SOURCES
# ============================================================
ws_sources = wb.create_sheet("Sources")

source_headers = ["Source ID", "Publication", "Publisher", "URL", "Data Year", "Date Accessed", "Notes"]
for col, h in enumerate(source_headers, 1):
    ws_sources.cell(row=1, column=col, value=h)
style_header(ws_sources, 1, len(source_headers))

sources = [
    # Coal sources
    ["C1", "China NBS Energy Production Statistics", "National Bureau of Statistics", "stats.gov.cn/english/PressRelease/202501/t20250124_1958444.html", "2024", "2026-03-07", ""],
    ["C2", "Qinhuangdao Coal Price Data", "CEIC Data / SunSirs", "ceicdata.com / sunsirs.com", "2024", "2026-03-07", "5500 kcal NAR"],
    ["C3", "IEA Coal Mid-Year Update 2025 - Prices", "IEA", "iea.org/reports/coal-mid-year-update-2025/prices", "2024", "2026-03-07", ""],
    ["C4", "Ministry of Coal Year End Review 2024", "Press Information Bureau, India", "pib.gov.in/PressReleasePage.aspx?PRID=2088467", "FY2024-25", "2026-03-07", ""],
    ["C5", "Coal India Ltd Notified Prices", "Coal India / Ministry of Coal", "coal.gov.in", "2024", "2026-03-07", "Last revision May 2023; +Rs 10/t Apr 2025"],
    ["C6", "Indonesia ESDM Coal Production Data", "Ministry of Energy & Mineral Resources", "minerba.esdm.go.id", "2024", "2026-03-07", "831-836 Mt record"],
    ["C7", "Indonesian HBA Reference Prices", "MEMR / Minerba", "minerba.esdm.go.id/harga_acuan", "2024", "2026-03-07", "Monthly HBA prices"],
    ["C8", "EIA Annual Coal Report 2024", "US Energy Information Administration", "eia.gov/coal/annual/", "2024", "2026-03-07", "512.5 MMst; prices by type"],
    ["C9", "Energy Institute Statistical Review 2025", "Energy Institute", "energyinst.org/statistical-review", "2024", "2026-03-07", "74th edition; comprehensive"],
    ["C10", "Australia DISR Resources & Energy Quarterly", "Geoscience Australia / DISR", "ga.gov.au/aecr2025/coal", "2024", "2026-03-07", ""],
    ["C11", "Russian Coal Industry 2024 Analysis", "The Coal Trader / The Coal Hub", "thecoaltrader.com / thecoalhub.com", "2024", "2026-03-07", "427 Mt; -2.5%"],
    ["C12", "Richards Bay Coal Terminal Data", "S&P Global Commodity Insights", "spglobal.com/commodity-insights", "2024", "2026-03-07", "52.1 Mt exported"],
    ["C13", "IEA Coal 2024 Report", "IEA", "iea.org/reports/coal-2024", "2024", "2026-03-07", "Comprehensive global coal analysis"],
    ["C14", "EURACOAL Market Report 2025", "EURACOAL", "euracoal.eu", "2024", "2026-03-07", "108.5 Mt Kazakhstan"],
    ["C15", "Mongolia Coal Production 2024", "Mysteel / CEIC", "mysteel.net / ceicdata.com", "2024", "2026-03-07", "98 Mt record; +20%"],
    ["C16", "Eurostat Coal Statistics", "Eurostat", "ec.europa.eu/eurostat", "2024", "2026-03-07", "EU coal production data"],
    ["C17", "Turkey Coal Production", "CEIC Data", "ceicdata.com", "2024", "2026-03-07", "+17% growth"],
    ["C18", "Poland Coal Production", "Eurostat / EURACOAL", "euracoal.eu", "2024", "2026-03-07", "44 Mt hard + 40 Mt lignite"],
    ["C19", "Colombia Coal - Drummond Ltd", "Drummond", "drummondltd.com", "2024", "2026-03-07", ""],
    ["C20", "Vietnam Coal Production", "CEIC Data / The Shiv", "ceicdata.com", "2024", "2026-03-07", "43.8 Mt"],
    ["C21", "Canada Coal Facts", "Natural Resources Canada", "natural-resources.canada.ca", "2024", "2026-03-07", "42.6 Mt; 67% met"],
    ["C22", "Czech Republic Coal", "CEIC / Eurostat", "ceicdata.com / ec.europa.eu", "2024", "2026-03-07", "~25 Mt; -17%"],
    ["C23", "Ukraine Coal Production", "CEIC Data", "ceicdata.com", "2024", "2026-03-07", "17.4 Mt; war-impacted"],
    ["C24", "Pakistan Coal Production", "CEIC / GEM", "ceicdata.com / gem.wiki", "2023-24", "2026-03-07", "Thar lignite growing"],
    ["C25", "Mozambique Coal", "GEM / Coaltrans", "gem.wiki/Moatize_Coal_Mine", "2024", "2026-03-07", "16.3 Mt; +9.4%"],
    ["C26", "Philippines Coal - Semirara", "GMA News / PNA", "gmanetwork.com / pna.gov.ph", "2024", "2026-03-07", "Record shipments"],
    ["C27", "Botswana Coal - Morupule", "GEM / Mining Technology", "gem.wiki/Morupule_Coal_Mine", "2023-24", "2026-03-07", "~2-3 Mt; expansion to 10 Mtpa by 2027"],
    # Oil sources
    ["O1", "EIA Brent/WTI Spot Price History", "US EIA", "eia.gov/dnav/pet/hist/rbrtem.htm", "2023", "2026-03-07", ""],
    ["O2", "EIA International Energy Statistics", "US EIA", "eia.gov/international/data/world", "2023", "2026-03-07", "Crude + condensate by country"],
    ["O3", "Urals Crude Price Data", "Statista / S&P Global", "statista.com", "2023", "2026-03-07", "$61.79 annual avg"],
    ["O4", "Russian Oil Sanctions Analysis", "Dallas Fed", "dallasfed.org", "2023", "2026-03-07", ""],
    ["O5", "OPEC Annual Statistical Bulletin 2024", "OPEC", "opec.org/assets/assetdb/asb-2024.pdf", "2023", "2026-03-07", "13 basket grade prices"],
    ["O6", "Western Canadian Select Price", "Statista / Alberta Energy", "statista.com", "2023", "2026-03-07", "$58.97 annual avg"],
    ["O7", "Energy Institute Statistical Review 2024", "Energy Institute", "energyinst.org/statistical-review", "2023", "2026-03-07", "73rd edition"],
    ["O8", "Murban Crude (IFAD)", "ICE Futures Abu Dhabi", "theice.com", "2023", "2026-03-07", ""],
    ["O9", "Iran Oil Pricing Analysis", "S&P Global / Academic", "various", "2023", "2026-03-07", "Sanctions discount estimated"],
    ["O10", "Petrobras / Brazil Crude", "Petrobras / ANP", "petrobras.com.br", "2023", "2026-03-07", "Tupi/Lula pre-salt"],
    ["O11", "Pemex Petroleum Statistics", "Pemex", "pemex.com", "2023", "2026-03-07", "Maya export price"],
    ["O12", "CPC Blend / Kazakhstan Crude", "S&P Global", "spglobal.com", "2023", "2026-03-07", ""],
    ["O13", "Guyana Production Data", "EIA / Hess Corp", "eia.gov", "2023", "2026-03-07", "Fastest growing"],
    ["O14", "Colombia Crude Benchmarks", "Ecopetrol / ANH", "various", "2023", "2026-03-07", "Vasconia, Castilla"],
    ["O15", "Azerbaijan / BTC Pipeline", "SOCAR / BP", "various", "2023", "2026-03-07", "Azeri Light"],
    ["O16", "Ecuador Oriente Crude", "Petroecuador", "various", "2023", "2026-03-07", ""],
    # Gas sources
    ["G1", "Energy Institute Statistical Review 2025", "Energy Institute", "energyinst.org/statistical-review", "2024", "2026-03-07", "Production by country"],
    ["G2", "EIA Henry Hub Spot Prices", "US EIA", "eia.gov/dnav/ng/hist/rngwhhdm.htm", "2023-24", "2026-03-07", "$2.57 (2023), $2.21 (2024)"],
    ["G3", "Russia Gas Pricing", "IGU WGPS / Academic", "igu.org", "2023", "2026-03-07", "Regulated ~5500 RUB/tcm"],
    ["G4", "GCC Gas Prices", "CSIS / IGU", "csis.org/analysis/gulf-gas-projects-still-priority-despite-high-costs", "2023", "2026-03-07", "Administered prices"],
    ["G5", "China Gas Market", "CEIC / IEA", "ceicdata.com / iea.org", "2023", "2026-03-07", "Mixed pricing"],
    ["G6", "AECO-C Price", "Alberta Energy Regulator", "aer.ca", "2023-24", "2026-03-07", "C$1.45/GJ in 2024"],
    ["G7", "Qatar Gas Pricing", "IGU / QatarEnergy", "igu.org", "2023", "2026-03-07", ""],
    ["G8", "Australia LNG", "EI / ACCC", "energyinst.org", "2023", "2026-03-07", ""],
    ["G9", "TTF Gas Price", "EIA / ICE", "eia.gov", "2023-24", "2026-03-07", "$13.10 (2023), ~$10.37 (2024)"],
    ["G10", "Algeria Gas", "Sonatrach / IEA", "various", "2023", "2026-03-07", ""],
    ["G11", "Central Asia Gas Prices", "IGU / Academic", "igu.org", "2023", "2026-03-07", "Regulated/subsidized"],
    ["G12", "Malaysia LNG", "Petronas / IGU", "various", "2023", "2026-03-07", ""],
    ["G13", "Egypt Gas", "EGAS / IEA", "various", "2023", "2026-03-07", "Gradually deregulating"],
    ["G14", "Indonesia Gas", "SKK Migas / IEA", "various", "2023", "2026-03-07", ""],
    ["G15", "Argentina Gas", "ENARGAS / IEA", "various", "2023", "2026-03-07", "Vaca Muerta"],
    ["G16", "Nigeria Gas", "NNPC / IEA", "various", "2023", "2026-03-07", "Dual pricing"],
    ["G17", "UK NBP Price", "ICE / Statista", "various", "2023", "2026-03-07", "Tracks TTF"],
    ["G18", "Pakistan Gas", "OGRA / IEA", "various", "2023", "2026-03-07", ""],
    ["G19", "Azerbaijan Gas", "SOCAR / TANAP", "various", "2023", "2026-03-07", "SGC pipeline"],
    ["G20", "India Gas", "PPAC / MoPNG", "ppac.gov.in", "2023", "2026-03-07", "APM pricing"],
    ["G21", "Mexico Gas", "SENER / CNH", "various", "2023", "2026-03-07", "HH-linked"],
    ["G22", "Thailand Gas", "EPPO / PTT", "various", "2023", "2026-03-07", ""],
    ["G23", "Trinidad Gas", "NGC / Atlantic LNG", "various", "2023", "2026-03-07", ""],
    ["G24", "Bangladesh Gas", "Petrobangla", "various", "2023", "2026-03-07", ""],
    ["G25", "Israel Gas", "Delek / NewMed", "various", "2023", "2026-03-07", "Leviathan, Tamar"],
    ["G26", "Brazil Gas", "ANP / Petrobras", "various", "2023", "2026-03-07", "Pre-salt associated"],
    # Transport sources
    ["T1", "IEA Coal 2025 - Prices and Costs", "IEA", "iea.org/reports/coal-2025/prices-and-costs", "2024", "2026-03-07", "Freight charts"],
    ["T2", "IEA Coal 2023 PDF Report", "IEA", "iea.org", "2023", "2026-03-07", "Russian coal freight"],
    ["T3", "China Coal Logistics", "The Coal Trader / SunSirs", "thecoaltrader.com", "2024", "2026-03-07", "Pit-to-port costs"],
    ["T4", "Indian Railways Coal Transport", "IEEFA", "ieefa.org", "2023", "2026-03-07", "Coal heavy burden report"],
    ["T5", "EIA Coal Transportation Rates", "US EIA", "eia.gov/coal/transportationrates/", "2023", "2026-03-07", "Rail rates by basin"],
    ["T6", "Australia Coal Rail", "ARTC / industry", "various", "2024", "2026-03-07", "Hunter Valley"],
    ["T7", "Indonesia Barge Costs", "Breakwave Advisors", "breakwaveadvisors.com", "2024", "2026-03-07", "Kalimantan barge"],
    ["T8", "South Africa Rail (Transnet)", "S&P Global", "spglobal.com", "2024", "2026-03-07", ""],
    ["T9", "Russia Rail Costs", "The Coal Hub / Argus", "thecoalhub.com / argusmedia.com", "2024", "2026-03-07", "RZD tariff increases"],
    ["T10", "EIA Crude Oil Tanker Rates", "US EIA", "eia.gov/todayinenergy/detail.php?id=67064", "2023", "2026-03-07", ""],
    ["T11", "Russia Oil Rerouting Costs", "Argus / UNCTAD", "argusmedia.com / unctad.org", "2023", "2026-03-07", "Post-sanctions longer routes"],
    ["T12", "Oxford Energy LNG Shipping", "Oxford Institute for Energy Studies", "oxfordenergy.org", "2023", "2026-03-07", "NG-188 Shipping Chokepoints"],
    ["T13", "US LNG Cost Analysis", "CSIS / Timera Energy", "csis.org / timera-energy.com", "2023", "2026-03-07", "Full chain costs"],
    ["T14", "Qatar LNG Costs", "IGU / Academic", "igu.org", "2023", "2026-03-07", "Lowest landed cost"],
    ["T15", "Australia LNG Costs", "ACCC / FTI Consulting", "accc.gov.au", "2023", "2026-03-07", ""],
]

for r, row in enumerate(sources, 2):
    for c, val in enumerate(row, 1):
        ws_sources.cell(row=r, column=c, value=val).border = thin_border

auto_width(ws_sources)

# ============================================================
# SAVE
# ============================================================
output_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Task1_Global_Fossil_Fuel_Market_Value.xlsx")
wb.save(output_path)
print(f"Excel workbook saved to: {output_path}")
print(f"Sheets: {wb.sheetnames}")
